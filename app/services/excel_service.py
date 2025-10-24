"""
엑셀 다운로드 서비스 (xlsx)

openpyxl을 사용하여 다양한 데이터를 엑셀 파일로 변환
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)

# 한국도로공사 브랜드 컬러
HEADER_FILL = PatternFill(start_color="0A2986", end_color="0A2986", fill_type="solid")
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name='맑은 고딕', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExcelService:
    """엑셀 파일 생성 서비스"""

    @staticmethod
    def create_workbook_from_data(
        data: List[Dict[str, Any]],
        headers: Dict[str, str],
        sheet_name: str = "Sheet1",
        title: str = None
    ) -> BytesIO:
        """
        데이터를 엑셀 파일로 변환

        Args:
            data: 데이터 목록 (dict의 리스트)
            headers: 컬럼 헤더 매핑 {'field_key': '한글 헤더명'}
            sheet_name: 시트 이름
            title: 엑셀 파일 상단 타이틀 (선택)

        Returns:
            BytesIO: 엑셀 파일 바이너리 데이터
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        current_row = 1

        # 타이틀 추가 (선택사항)
        if title:
            ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(name='맑은 고딕', size=14, bold=True, color="0A2986")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            current_row = 2

        # 헤더 작성
        for col_idx, (field_key, header_name) in enumerate(headers.items(), start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header_name
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # 데이터 작성
        for row_idx, row_data in enumerate(data, start=current_row):
            for col_idx, field_key in enumerate(headers.keys(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(field_key, '')

                # 날짜/시간 포맷팅
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')

                # Boolean 변환
                if isinstance(value, bool):
                    value = 'O' if value else 'X'

                cell.value = value
                cell.font = CELL_FONT
                cell.border = THIN_BORDER

                # 숫자는 우측 정렬, 나머지는 좌측 정렬
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # 컬럼 너비 자동 조정
        for col_idx, field_key in enumerate(headers.keys(), start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(headers[field_key]) + 2  # 헤더 길이

            # 데이터 길이 확인
            for row_data in data[:100]:  # 최대 100개 행만 확인 (성능)
                value = str(row_data.get(field_key, ''))
                max_length = max(max_length, len(value) + 2)

            # 최대 너비 제한
            max_length = min(max_length, 50)
            ws.column_dimensions[column_letter].width = max_length

        # BytesIO로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel file created: {len(data)} rows, {len(headers)} columns")
        return output

    @staticmethod
    def create_conversations_excel(conversations: List[Dict]) -> BytesIO:
        """
        대화 내역 엑셀 생성

        Args:
            conversations: 대화 내역 목록

        Returns:
            BytesIO: 엑셀 파일
        """
        headers = {
            'id': 'ID',
            'user_id': '사용자 ID',
            'user_name': '사용자명',
            'question': '질문',
            'response': '응답',
            'model_name': '모델',
            'response_time': '응답시간(ms)',
            'token_count': '토큰 수',
            'created_at': '생성일시',
            'satisfaction_rating': '만족도'
        }

        return ExcelService.create_workbook_from_data(
            data=conversations,
            headers=headers,
            sheet_name="대화내역",
            title="ex-GPT 대화 내역"
        )

    @staticmethod
    def create_usage_statistics_excel(stats: List[Dict]) -> BytesIO:
        """
        사용 통계 엑셀 생성

        Args:
            stats: 통계 데이터 목록

        Returns:
            BytesIO: 엑셀 파일
        """
        headers = {
            'date': '날짜',
            'count': '질문 수',
            'unique_users': '사용자 수',
            'avg_response_time': '평균 응답시간(ms)',
            'satisfaction_avg': '평균 만족도'
        }

        return ExcelService.create_workbook_from_data(
            data=stats,
            headers=headers,
            sheet_name="사용통계",
            title="ex-GPT 사용 통계"
        )

    @staticmethod
    def create_satisfaction_excel(surveys: List[Dict]) -> BytesIO:
        """
        만족도 조사 엑셀 생성

        Args:
            surveys: 만족도 조사 목록

        Returns:
            BytesIO: 엑셀 파일
        """
        headers = {
            'id': 'ID',
            'conversation_id': '대화 ID',
            'user_id': '사용자 ID',
            'rating': '만족도(1-5)',
            'feedback': '피드백',
            'created_at': '작성일시'
        }

        return ExcelService.create_workbook_from_data(
            data=surveys,
            headers=headers,
            sheet_name="만족도조사",
            title="ex-GPT 만족도 조사 결과"
        )

    @staticmethod
    def create_dictionaries_excel(dictionaries: List[Dict]) -> BytesIO:
        """
        사전 목록 엑셀 생성

        Args:
            dictionaries: 사전 목록

        Returns:
            BytesIO: 엑셀 파일
        """
        headers = {
            'dict_id': '사전 ID',
            'dict_name': '사전명',
            'dict_type': '사전 종류',
            'dict_desc': '설명',
            'word_count': '용어 수',
            'case_sensitive': '대소문자 구분',
            'word_boundary': '띄어쓰기 구분',
            'use_yn': '사용 여부',
            'created_at': '생성일시'
        }

        return ExcelService.create_workbook_from_data(
            data=dictionaries,
            headers=headers,
            sheet_name="사전목록",
            title="동의어 사전 목록"
        )

    @staticmethod
    def create_dictionary_terms_excel(terms: List[Dict], dict_name: str = None) -> BytesIO:
        """
        사전 용어 엑셀 생성

        Args:
            terms: 용어 목록
            dict_name: 사전명 (타이틀용)

        Returns:
            BytesIO: 엑셀 파일
        """
        headers = {
            'term_id': '용어 ID',
            'main_term': '대표 용어',
            'main_alias': '표준 표기',
            'synonym1': '동의어1',
            'synonym2': '동의어2',
            'synonym3': '동의어3',
            'category': '카테고리',
            'definition': '정의',
            'use_yn': '사용 여부',
            'created_at': '생성일시'
        }

        title = f"{dict_name} - 용어 목록" if dict_name else "사전 용어 목록"

        return ExcelService.create_workbook_from_data(
            data=terms,
            headers=headers,
            sheet_name="용어목록",
            title=title
        )

    @staticmethod
    def create_users_excel(users: List[Dict]) -> BytesIO:
        """
        사용자 목록 엑셀 생성

        Args:
            users: 사용자 목록

        Returns:
            BytesIO: 엑셀 파일
        """
        headers = {
            'id': 'ID',
            'username': '사용자명',
            'full_name': '성명',
            'email': '이메일',
            'department': '부서',
            'is_active': '활성 여부',
            'created_at': '가입일시',
            'last_login': '마지막 로그인'
        }

        return ExcelService.create_workbook_from_data(
            data=users,
            headers=headers,
            sheet_name="사용자목록",
            title="ex-GPT 사용자 목록"
        )

    @staticmethod
    def create_notices_excel(notices: List[Dict]) -> BytesIO:
        """
        공지사항 목록 엑셀 생성

        Args:
            notices: 공지사항 목록

        Returns:
            BytesIO: 엑셀 파일
        """
        headers = {
            'id': 'ID',
            'title': '제목',
            'content': '내용',
            'author': '작성자',
            'is_pinned': '상단고정',
            'view_count': '조회수',
            'created_at': '작성일시',
            'updated_at': '수정일시'
        }

        return ExcelService.create_workbook_from_data(
            data=notices,
            headers=headers,
            sheet_name="공지사항",
            title="ex-GPT 공지사항 목록"
        )

    @staticmethod
    def create_stt_batches_excel(batches: List[Dict]) -> BytesIO:
        """
        STT 배치 목록 엑셀 생성

        Args:
            batches: STT 배치 목록

        Returns:
            BytesIO: 엑셀 파일
        """
        headers = {
            'id': 'ID',
            'batch_name': '배치명',
            'status': '상태',
            'total_files': '총 파일 수',
            'completed_files': '완료 파일 수',
            'failed_files': '실패 파일 수',
            'created_by': '생성자',
            'created_at': '생성일시',
            'completed_at': '완료일시'
        }

        return ExcelService.create_workbook_from_data(
            data=batches,
            headers=headers,
            sheet_name="STT배치",
            title="STT 음성 전사 배치 목록"
        )


# 편의 함수
def export_to_excel(
    data: List[Dict],
    headers: Dict[str, str],
    filename: str = "export.xlsx",
    sheet_name: str = "Sheet1",
    title: str = None
) -> BytesIO:
    """
    데이터를 엑셀로 내보내기 (편의 함수)

    Args:
        data: 데이터 목록
        headers: 헤더 매핑
        filename: 파일명 (사용되지 않음, 호환성을 위해 유지)
        sheet_name: 시트명
        title: 타이틀

    Returns:
        BytesIO: 엑셀 파일
    """
    return ExcelService.create_workbook_from_data(
        data=data,
        headers=headers,
        sheet_name=sheet_name,
        title=title
    )
